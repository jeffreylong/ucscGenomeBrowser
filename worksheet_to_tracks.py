'''
Created on Aug 19, 2016

@author: jaykasberger
'''
import argparse
import sys
import re
from openpyxl.reader.excel import load_workbook  # @UnresolvedImport

def main(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True, help='Excel file to import')
    parser.add_argument('--output', required=True, help='Track hub file to write')
    args = parser.parse_args()
    input_filename = args.input
    input_workbook = load_workbook(input_filename)    
    
    track_worksheet = input_workbook.get_sheet_by_name("Sheet1")
    trackDb_txt = open(args.output, "w")
    row = 2
    while track_worksheet.cell(row=row, column=1).value:
        track = track_worksheet.cell(row=row, column=1).value
        bigDataUrl = track_worksheet.cell(row=row, column=2).value
        shortLabel = track_worksheet.cell(row=row, column=3).value
        longLabel = track_worksheet.cell(row=row, column=4).value
        file_type = track_worksheet.cell(row=row, column=5).value

        trackDb_txt.write("track " + track + "\n")
        trackDb_txt.write("bigDataUrl " + bigDataUrl + "\n")
        trackDb_txt.write("shortLabel " + shortLabel +  "\n")
        trackDb_txt.write("longLabel " + longLabel + "\n")
        trackDb_txt.write("type " + file_type + "\n")
        trackDb_txt.write("\n")
        row = row + 1
    
    trackDb_txt.close()


    
if __name__ == '__main__':
    main(sys.argv[1:])
